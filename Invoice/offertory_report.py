"""
Church Offertory Report Generator
Generates offertory collection reports based on receipt data and Excel templates.
"""

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None
    load_workbook = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

import os
import json
import datetime
from pathlib import Path
import glob

class OffertoryReportGenerator:
    def __init__(self, template_path=None):
        """Initialize the report generator with template path"""
        if template_path is None:
            # Default to template in project root
            script_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(script_dir)
            template_path = os.path.join(project_root, "Offertory_Report_Template.xlsx")
        
        self.template_path = template_path
        self.output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Receipts store", "Reports")
        os.makedirs(self.output_dir, exist_ok=True)
    
    def collect_receipt_data(self, date_filter=None):
        """
        Collect receipt data from all generated receipts
        Args:
            date_filter: Optional date string (YYYY-MM-DD) to filter receipts by date
        Returns:
            List of receipt data dictionaries
        """
        receipts_data = []
        
        # Look for receipt data in multiple locations
        search_paths = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads"),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "Receipts store"),
        ]
        
        for search_path in search_paths:
            if os.path.exists(search_path):
                # Find all JSON result files
                json_files = glob.glob(os.path.join(search_path, "*_result.json"))
                
                for json_file in json_files:
                    try:
                        with open(json_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                        
                        # Extract receipt data from the analysis result
                        receipt_data = self._extract_receipt_data_from_result(data)
                        
                        # Apply date filter if specified
                        if date_filter:
                            receipt_date = receipt_data.get('InvoiceDate', '')
                            if not receipt_date.startswith(date_filter):
                                continue
                        
                        receipts_data.append(receipt_data)
                    
                    except Exception as e:
                        print(f"Error reading {json_file}: {e}")
                        continue
        
        return receipts_data
    
    def _extract_receipt_data_from_result(self, result_data):
        """Extract receipt data from JSON result file"""
        receipt_data = {}
        
        # Try to extract from the result structure
        if 'result' in result_data and 'contents' in result_data['result']:
            contents = result_data['result']['contents']
            if contents and 'fields' in contents[0]:
                fields = contents[0]['fields']
                
                # Extract all the fields we need
                field_mappings = {
                    'InvoiceDate': ['InvoiceDate'],
                    'Name': ['Name'],
                    'Address': ['Address'],
                    'MobileNumber': ['MobileNumber'],
                    'TitheMonth': ['TitheMonth'],
                    'TitheAmount': ['TitheAmount'],
                    'MembershipMonth': ['MembershipMonth'],
                    'MembershipAmount': ['MembershipAmount'],
                    'BirthdayThankOffering': ['BirthdayThankOffering'],
                    'WeddingAnniversaryThankOffering': ['WeddingAnniversaryThankOffering'],
                    'HomeMissionPledges': ['HomeMissionPledges'],
                    'MissionAndEvangelismFund': ['MissionAndEvangelismFund'],
                    'StStephensSocialAidFund': ['StStephensSocialAidFund'],
                    'DonationFor': ['DonationFor'],
                    'DonationAmount': ['DonationAmount'],
                    'SpecialThanksAmount': ['SpecialThanksAmount'],
                    'CharityFundAmount': ['CharityFundAmount'],
                    'HarvestAuctionComment': ['HarvestAuctionComment'],
                    'HarvestAuctionAmount': ['HarvestAuctionAmount'],
                    'OnlineChequeNo': ['OnlineChequeNo']
                }
                
                for field_name, possible_keys in field_mappings.items():
                    for key in possible_keys:
                        if key in fields:
                            field_data = fields[key]
                            value = (
                                field_data.get('value') or
                                field_data.get('valueDate') or
                                field_data.get('valueString') or
                                field_data.get('valueNumber') or
                                field_data.get('valueInteger')
                            )
                            if value:
                                receipt_data[field_name] = value
                            break
        
        return receipt_data
    
    def generate_offertory_report(self, service_date=None, service_type="Worship Service"):
        """
        Generate offertory report similar to the sample image
        Args:
            service_date: Date of the service (YYYY-MM-DD format)
            service_type: Type of service (default: "Worship Service")
        Returns:
            Path to generated Excel report
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel report generation. Please install it with: pip install openpyxl")
        
        if service_date is None:
            service_date = datetime.date.today().strftime("%Y-%m-%d")
        
        # Collect receipt data for the specified date
        receipts_data = self.collect_receipt_data(service_date)
        
        # Load the template
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        wb = load_workbook(self.template_path)
        ws = wb.active
        
        # Update header information
        self._update_header_info(ws, service_date, service_type)
        
        # Fill Section A - Bag Offertory Collection (Cash denominations)
        total_cash_a = self._fill_section_a(ws, receipts_data)
        
        # Fill Section B - Special Offertory (Individual contributions)
        total_special_b = self._fill_section_b(ws, receipts_data)
        
        # Update totals
        self._update_totals(ws, total_cash_a, total_special_b)
        
        # Generate filename and save
        report_filename = f"Offertory_Report_{service_date}_{service_type.replace(' ', '_')}.xlsx"
        report_path = os.path.join(self.output_dir, report_filename)
        
        wb.save(report_path)
        return report_path
    
    def generate_offertory_report_pdf(self, service_date=None, service_type="Worship Service"):
        """
        Generate offertory report as PDF based on Excel template structure
        Args:
            service_date: Date of the service (YYYY-MM-DD format)
            service_type: Type of service (default: "Worship Service")
        Returns:
            Path to generated PDF report
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF report generation. Please install it with: pip install reportlab")
        
        if service_date is None:
            service_date = datetime.date.today().strftime("%Y-%m-%d")
        
        # Collect receipt data for the specified date
        receipts_data = self.collect_receipt_data(service_date)
        
        # First, read the Excel template to understand its structure
        template_structure = self._read_excel_template()
        
        # Generate filename
        report_filename = f"Offertory_Report_{service_date}_{service_type.replace(' ', '_')}.pdf"
        report_path = os.path.join(self.output_dir, report_filename)
        
        # Create PDF document with same layout as Excel template
        doc = SimpleDocTemplate(report_path, pagesize=A4, 
                               topMargin=0.5*inch, bottomMargin=0.5*inch,
                               leftMargin=0.5*inch, rightMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Add custom styles to match Excel formatting
        custom_title_style = styles['Title'].clone('CustomTitle')
        custom_title_style.fontSize = 16
        custom_title_style.spaceAfter = 10
        custom_title_style.alignment = 1  # Center alignment
        
        # Title (from template)
        title = Paragraph("<b>Offertory (Cash and Cheque) Collection Record</b>", custom_title_style)
        elements.append(title)
        elements.append(Spacer(1, 15))
        
        # Date and Service Info section (matching Excel template layout)
        date_obj = datetime.datetime.strptime(service_date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d/%m/%y")
        
        # Create header info table to match Excel template
        header_info = [
            ['', '', '', '', ''],
            ['Date:', formatted_date, '', 'Sunday:', service_type],
            ['', '', '', '', '']
        ]
        
        header_table = Table(header_info, colWidths=[1*inch, 1.2*inch, 1*inch, 0.8*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,1), (1,1), 'LEFT'),    # Date section
            ('ALIGN', (3,1), (4,1), 'LEFT'),    # Sunday section
            ('FONTNAME', (0,1), (0,1), 'Helvetica-Bold'),  # "Date:" label
            ('FONTNAME', (3,1), (3,1), 'Helvetica-Bold'),  # "Sunday:" label
            ('BOX', (0,1), (4,1), 1, colors.black),
            ('INNERGRID', (0,1), (4,1), 0.5, colors.black),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Section A - Bag Offertory Collection (matching Excel template)
        section_a_title = Paragraph("<b>A. Bag Offertory Collection - Cash</b>", styles['Heading3'])
        elements.append(section_a_title)
        elements.append(Spacer(1, 10))
        
        # Calculate total cash from receipts
        total_cash = self._calculate_total_cash(receipts_data)
        
        # Cash denominations table (exactly as in Excel template)
        cash_data = [
            ['Denomination', 'No. of Notes/Coins', 'Amount (₹)'],
            ['2000', '', ''],
            ['500', '', ''],
            ['200', '', ''],
            ['100', '', ''],
            ['50', '', ''],
            ['20', '', ''],
            ['10', '', ''],
            ['5', '', ''],
            ['2', '', ''],
            ['1', '', ''],
            ['', '', ''],
            ['TOTAL CASH - A', '', f'{total_cash:.2f}'],
        ]
        
        # Style the cash table to match Excel appearance
        cash_table = Table(cash_data, colWidths=[2*inch, 2*inch, 2*inch])
        cash_table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            
            # Grid and borders
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BOX', (0,0), (-1,-1), 2, colors.black),
            
            # Total row styling (last row)
            ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            
            # Padding
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(cash_table)
        elements.append(Spacer(1, 25))
        
        # Section B - Special Offertory (matching Excel template)
        section_b_title = Paragraph("<b>B. Special Offertory - Cash(B1) & Cheque(B2)</b>", styles['Heading3'])
        elements.append(section_b_title)
        elements.append(Spacer(1, 10))
        
        # Create special offertory table with data
        special_data, total_special = self._create_special_offertory_data(receipts_data)
        
        # Style the special offertory table to match Excel
        special_table = Table(special_data, colWidths=[0.8*inch, 2.2*inch, 2.2*inch, 1*inch])
        special_table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            
            # Grid and borders
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BOX', (0,0), (-1,-1), 2, colors.black),
            
            # Total row styling (last row)
            ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('SPAN', (0,-1), (2,-1)),  # Merge first three columns for total
            
            # Data rows - left align name and description, center others
            ('ALIGN', (1,1), (1,-2), 'LEFT'),    # Names column
            ('ALIGN', (2,1), (2,-2), 'LEFT'),    # Description column
            
            # Padding
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(special_table)
        elements.append(Spacer(1, 25))
        
        # Grand Total (matching Excel template)
        grand_total = total_cash + total_special
        grand_total_data = [
            ['GRAND TOTAL (A + B)', f'{grand_total:.2f}']
        ]
        
        grand_total_table = Table(grand_total_data, colWidths=[4.5*inch, 1.5*inch])
        grand_total_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 14),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOX', (0,0), (-1,-1), 2, colors.black),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ]))
        elements.append(grand_total_table)
        
        # Add signature section (as in Excel template)
        elements.append(Spacer(1, 40))
        signature_data = [
            ['', '', ''],
            ['Prepared by:', 'Checked by:', 'Approved by:'],
            ['', '', ''],
            ['_________________', '_________________', '_________________'],
            ['Signature & Date', 'Signature & Date', 'Signature & Date']
        ]
        
        signature_table = Table(signature_data, colWidths=[2*inch, 2*inch, 2*inch])
        signature_table.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTNAME', (0,4), (-1,4), 'Helvetica-Oblique'),
            ('FONTSIZE', (0,4), (-1,4), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(signature_table)
        
        # Build PDF
        doc.build(elements)
        return report_path
    
    def _read_excel_template(self):
        """Read the Excel template to understand its structure"""
        template_structure = {
            'title': 'Offertory (Cash and Cheque) Collection Record',
            'sections': {
                'header': ['Date:', 'Sunday:'],
                'section_a': 'A. Bag Offertory Collection - Cash',
                'section_b': 'B. Special Offertory - Cash(B1) & Cheque(B2)',
                'denominations': ['2000', '500', '200', '100', '50', '20', '10', '5', '2', '1']
            }
        }
        return template_structure
    
    def _calculate_total_cash(self, receipts_data):
        """Calculate total cash amount from receipts"""
        total_cash = 0
        for receipt in receipts_data:
            payment_method = receipt.get('PaymentMethod', 'CASH')
            if payment_method.upper() == 'CASH':
                for field in ['TitheAmount', 'MembershipAmount', 'BirthdayThankOffering',
                             'WeddingAnniversaryThankOffering', 'SpecialThanksAmount',
                             'MissionAndEvangelismFund', 'CharityFundAmount',
                             'StStephensSocialAidFund', 'HarvestAuctionAmount',
                             'DonationAmount']:
                    amount = receipt.get(field, 0)
                    if amount:
                        try:
                            total_cash += float(amount)
                        except (ValueError, TypeError):
                            continue
        return total_cash
    
    def _create_special_offertory_data(self, receipts_data):
        """Create special offertory table data matching Excel template"""
        special_data = [['Sr.No.', 'Name', 'Contribution Details', 'Amount (₹)']]
        
        total_special = 0
        sr_no = 1
        
        # Process each receipt
        for receipt in receipts_data:
            name = receipt.get('Name', 'Unknown')
            if not name or name == 'Unknown':
                continue
            
            # Contribution types mapping (matching Excel template)
            contribution_types = {
                'TitheAmount': 'Tithe',
                'MembershipAmount': 'Membership',
                'BirthdayThankOffering': 'Birthday Thank Offering',
                'WeddingAnniversaryThankOffering': 'Wedding Anniversary Thank Offering',
                'SpecialThanksAmount': 'Special Thanks',
                'MissionAndEvangelismFund': 'Mission and Evangelism Fund',
                'CharityFundAmount': 'Charity Fund',
                'StStephensSocialAidFund': 'St.Stephens Social Aid Fund',
                'HarvestAuctionAmount': 'Harvest Auction',
                'DonationAmount': f"Donation for {receipt.get('DonationFor', 'General')}"
            }
            
            # Add each contribution as a separate row
            for field, description in contribution_types.items():
                amount = receipt.get(field, 0)
                if amount:
                    try:
                        amount_float = float(amount)
                        if amount_float > 0:
                            special_data.append([
                                str(sr_no), 
                                name[:25],  # Limit name length
                                description, 
                                f'{amount_float:.2f}'
                            ])
                            total_special += amount_float
                            sr_no += 1
                    except (ValueError, TypeError):
                        continue
        
        # Add empty rows if needed (to match Excel template spacing)
        while len(special_data) < 15:  # Ensure minimum rows for consistency
            special_data.append(['', '', '', ''])
        
        # Add total row
        special_data.append(['', 'TOTAL SPECIAL OFFERTORY - B', '', f'{total_special:.2f}'])
        
        return special_data, total_special
    
    def _update_header_info(self, ws, service_date, service_type):
        """Update the header information in the report"""
        # Find and update the date field
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and "Date:" in str(cell.value):
                    # Update date in adjacent cell
                    date_obj = datetime.datetime.strptime(service_date, "%Y-%m-%d")
                    formatted_date = date_obj.strftime("%d/%m/%y")
                    ws[f"{chr(ord(cell.column_letter) + 1)}{cell.row}"] = formatted_date
                
                if cell.value and "Sunday" in str(cell.value):
                    # Update service type
                    ws[cell.coordinate] = f"Sunday - {service_type}"
    
    def _fill_section_a(self, ws, receipts_data):
        """Fill Section A - Bag Offertory Collection with cash denominations"""
        # This section would be manually filled or based on cash collection data
        # For now, we'll calculate totals from receipts and estimate denominations
        
        total_cash = 0
        for receipt in receipts_data:
            # Sum all cash amounts from receipts
            payment_method = receipt.get('PaymentMethod', 'CASH')
            if payment_method.upper() == 'CASH':
                for field in ['TitheAmount', 'MembershipAmount', 'BirthdayThankOffering',
                             'WeddingAnniversaryThankOffering', 'SpecialThanksAmount',
                             'MissionAndEvangelismFund', 'CharityFundAmount',
                             'StStephensSocialAidFund', 'HarvestAuctionAmount',
                             'DonationAmount']:
                    amount = receipt.get(field, 0)
                    if amount:
                        try:
                            total_cash += float(amount)
                        except (ValueError, TypeError):
                            continue
        
        # Find the total cash cell in Section A and update it
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "TOTAL CASH - A" in str(cell.value):
                    # Update the adjacent cell with total
                    ws[f"{chr(ord(cell.column_letter) + 1)}{cell.row}"] = total_cash
                    break
        
        return total_cash
    
    def _fill_section_b(self, ws, receipts_data):
        """Fill Section B - Special Offertory with individual contributions"""
        total_special = 0
        sr_no = 1
        
        # Find the starting row for Section B data
        start_row = None
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value and "Special Offertory" in str(cell.value):
                    start_row = row_idx + 3  # Start a few rows below the header
                    break
            if start_row:
                break
        
        if not start_row:
            start_row = 15  # Default starting row if not found
        
        current_row = start_row
        
        for receipt in receipts_data:
            name = receipt.get('Name', 'Unknown')
            if not name or name == 'Unknown':
                continue
            
            # Calculate total amount for this person
            person_total = 0
            contributions = []
            
            # Check all contribution types
            contribution_types = {
                'TitheAmount': 'Tithe',
                'MembershipAmount': 'Membership',
                'BirthdayThankOffering': 'Birthday Offering',
                'WeddingAnniversaryThankOffering': 'Anniversary Offering',
                'SpecialThanksAmount': 'Special Thanks',
                'MissionAndEvangelismFund': 'Mission & Evangelism',
                'CharityFundAmount': 'Charity Fund',
                'StStephensSocialAidFund': 'Social Aid',
                'HarvestAuctionAmount': 'Harvest Auction',
                'DonationAmount': 'Donation'
            }
            
            for field, description in contribution_types.items():
                amount = receipt.get(field, 0)
                if amount:
                    try:
                        amount_val = float(amount)
                        if amount_val > 0:
                            person_total += amount_val
                            
                            # Add specific details
                            if field == 'TitheAmount' and receipt.get('TitheMonth'):
                                contributions.append(f"{description} ({receipt.get('TitheMonth')})")
                            elif field == 'MembershipAmount' and receipt.get('MembershipMonth'):
                                contributions.append(f"{description} ({receipt.get('MembershipMonth')})")
                            elif field == 'DonationAmount' and receipt.get('DonationFor'):
                                contributions.append(f"{description} ({receipt.get('DonationFor')})")
                            elif field == 'HarvestAuctionAmount' and receipt.get('HarvestAuctionComment'):
                                contributions.append(f"{description} ({receipt.get('HarvestAuctionComment')})")
                            else:
                                contributions.append(description)
                    except (ValueError, TypeError):
                        continue
            
            if person_total > 0 and contributions:
                # Fill the row with data
                # Assuming columns: Sr No, Name, Amount, Details
                ws[f"A{current_row}"] = sr_no  # Sr No
                ws[f"B{current_row}"] = name   # Name
                ws[f"C{current_row}"] = person_total  # Amount
                ws[f"D{current_row}"] = "; ".join(contributions)  # Details
                
                total_special += person_total
                sr_no += 1
                current_row += 1
        
        # Update the total for Section B
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "& B(I)" in str(cell.value):
                    # This is likely the total cell for Section B
                    # Find the amount cell in this row
                    amount_cell = None
                    for c in row:
                        if c.column > cell.column:
                            try:
                                if c.value is None or isinstance(c.value, (int, float)):
                                    amount_cell = c
                                    break
                            except:
                                continue
                    if amount_cell:
                        amount_cell.value = total_special
                    break
        
        return total_special
    
    def _update_totals(self, ws, total_cash_a, total_special_b):
        """Update the grand totals in the report"""
        grand_total = total_cash_a + total_special_b
        
        # Find and update grand total cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "GRAND TOTAL" in str(cell.value):
                    # Look for amount cells in this row
                    for c in row:
                        if c.column > cell.column:
                            try:
                                if c.value is None or isinstance(c.value, (int, float)):
                                    c.value = grand_total
                                    break
                            except:
                                continue
                    break
    
    def get_available_receipt_dates(self):
        """Get list of available receipt dates for report generation"""
        receipts_data = self.collect_receipt_data()
        dates = set()
        
        for receipt in receipts_data:
            date = receipt.get('InvoiceDate')
            if date:
                # Extract date part only
                if 'T' in date:
                    date = date.split('T')[0]
                dates.add(date)
        
        return sorted(list(dates))
    
    def generate_summary_report(self, start_date=None, end_date=None):
        """Generate a summary report for a date range"""
        if start_date is None:
            start_date = datetime.date.today().replace(day=1).strftime("%Y-%m-%d")
        if end_date is None:
            end_date = datetime.date.today().strftime("%Y-%m-%d")
        
        receipts_data = self.collect_receipt_data()
        
        # Filter by date range
        filtered_data = []
        for receipt in receipts_data:
            receipt_date = receipt.get('InvoiceDate', '')
            if 'T' in receipt_date:
                receipt_date = receipt_date.split('T')[0]
            
            if start_date <= receipt_date <= end_date:
                filtered_data.append(receipt)
        
        # Generate summary statistics
        summary = {
            'total_receipts': len(filtered_data),
            'total_amount': 0,
            'by_type': {},
            'by_date': {},
            'start_date': start_date,
            'end_date': end_date
        }
        
        for receipt in filtered_data:
            # Sum amounts
            receipt_total = 0
            for field in ['TitheAmount', 'MembershipAmount', 'BirthdayThankOffering',
                         'WeddingAnniversaryThankOffering', 'SpecialThanksAmount',
                         'MissionAndEvangelismFund', 'CharityFundAmount',
                         'StStephensSocialAidFund', 'HarvestAuctionAmount',
                         'DonationAmount']:
                amount = receipt.get(field, 0)
                if amount:
                    try:
                        receipt_total += float(amount)
                    except (ValueError, TypeError):
                        continue
            
            summary['total_amount'] += receipt_total
            
            # Group by date
            receipt_date = receipt.get('InvoiceDate', '')
            if 'T' in receipt_date:
                receipt_date = receipt_date.split('T')[0]
            
            if receipt_date not in summary['by_date']:
                summary['by_date'][receipt_date] = {'count': 0, 'amount': 0}
            
            summary['by_date'][receipt_date]['count'] += 1
            summary['by_date'][receipt_date]['amount'] += receipt_total
        
        return summary

def create_offertory_report(service_date=None, service_type="Worship Service"):
    """Convenience function to generate offertory report"""
    generator = OffertoryReportGenerator()
    return generator.generate_offertory_report(service_date, service_type)

def get_receipt_summary(start_date=None, end_date=None):
    """Convenience function to get receipt summary"""
    generator = OffertoryReportGenerator()
    return generator.generate_summary_report(start_date, end_date)
